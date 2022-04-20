# all imports needed for the project
import datetime
import os
import win32com.client
import pandas as pd
import openpyxl
import csv
import shutil


# This function converts the xlsx data into a csv per the requirement for the patron records folder
def convert_xlsx_csv():
    read_file = pd.read_excel("ProspectorPatrons.xlsx")
    read_file.to_csv("ProspectorPatrons.csv", index=None, header=True)


# This function allows us to delete all empty rows from the file as well in order to check that there are no empty
# new lines
def delete_rows(sheet):
    # iterate the sheet by rows
    for row in sheet.iter_rows():
        # all() return False if all of the row value is None
        if not all(cell.value for cell in row):
            # detele the empty row
            sheet.delete_rows(row[0].row, 1)
            # recursively call the remove() with modified sheet data
            delete_rows(sheet)
            return


# This function allows us to save the attachment directly from the email as long as outlook is installed
# and Libsys is a shared mailbox accessible on the computer. I am able to do this by starting a process of outlook
# in the code and then iterating to the most recent email with the attachment we are looking for.
def save_attachments(subject, messages):
    # check variable to ensure we find the email and return an error if not
    check = 0
    # for loop to iterate over all messages until we find the one with the correct attachment
    for message in messages:
        # luckily the message we want will always have the same subject allowing me to set a constant variable for it
        if message.Subject == subject:
            # using the win32 library and the objects and methods that come with it to save the attachment
            attachments = message.Attachments
            attachment = attachments.Item(2)
            attachment.SaveAsFile(os.path.join(path, str(attachment)))
            # marking it as read if needed
            if message.Subject == subject and message.Unread:
                message.Unread = False
            # loading a workbook in order to delete the rows and not have to worry about empty rows in the data
            xfile = openpyxl.load_workbook("ProspectorPatrons.xlsx")
            sheet = xfile["Sheet1"]
            delete_rows(sheet)
            for row in sheet.iter_rows():
                sheet.delete_rows(row[0].row, 1)
                break
            # saving the workbook to the proper xlsx file and converting it to a csv file
            xfile.save("ProspectorPatrons.xlsx")
            convert_xlsx_csv()
            check = 1
            break
    # checking for errors if the email wasn't found
    if check == 1:
        print("Email found and converted successfully")
    else:
        print("Correct email not found")
        return -1


def change_filename(filename1, filename2):
    # using the shutil library to move the file onto the Y: drive as the os library does not allow access across drives
    shutil.move(filename1, xlsx_file)
    shutil.move(filename2, csv_file)


def check_csv(csv_file):
    # checking the csv file for any errors in the UIDS so that they all end in 2341
    with open(csv_file) as csvfile:
        fileread = csv.reader(csvfile)
        for row in fileread:
            if row[0][-4:] == "2341":
                continue
            else:
                print("error in csv file")
                return -1
        print("File is correct")


# setting our default path variable to originally save the xlsx file to in order to manipulate it within the program
path = os.path.expanduser("C:\\Users\\garrettthompson_a\\Downloads\\Prospector\\ProspectorAutomatedProject")
# path to where we need the files to be
path2 = "Y:\\LB\\SharedSpace\\Systems\\Discovery\\Prospector\\"  # Patron Records\\
# finding the date for when we save the file
today = datetime.date.today()
# opening the outlook application and finding the proper inbox
outlook = win32com.client.Dispatch('outlook.application').GetNamespace("MAPI").Folders
folder = outlook(1)
inbox = folder.Folders("Inbox")
messages = inbox.Items
# sorting the messages because for some reason it doesn't always collect them in order of date and time
messages.Sort("[ReceivedTime]", True)
date = datetime.date.today().strftime('%Y-%m-%d')
# changing the file names to what we actually need them to be
xlsx_file = path2 + "minespatrons" + date + ".xlsx"
csv_file = path2 + "minespatrons" + date + ".csv"

# 3 function calls to actually perform everything
save_attachments("ProspectorPatrons", messages)
change_filename("ProspectorPatrons.xlsx", "ProspectorPatrons.csv")
check_csv(csv_file)
