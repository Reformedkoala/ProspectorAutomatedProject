# Pseudocode
# Download file from email and place in directory with project (Look at interacting directly with outlook) - start interacting with the shared box
# Open xlsx file and delete first three rows of data and check if last row has data not needed - done
# Save edited file as both csv and xlsx and add the date in the format YYYY-MM-DD to the end of the name before the extension
# Verify that all Unique Identifiers have 2341 as the last 4 digits of column 1 in the CSV file
# Move new files to //LB/SharedSpace/Systems/Discovery/Prospector/Patron Records/
# Delete the active file from the directory

import datetime
import os
import win32com.client
import pandas as pd
import openpyxl
import csv
import shutil


# This function converts the csv data that we have cleaned into something that we can use a bit
# easier and make it look like the data that we already have
def convert_xlsx_csv():
    read_file = pd.read_excel("ProspectorPatrons.xlsx")
    read_file.to_csv("ProspectorPatrons.csv", index=None, header=True)


# This function allows us to automatically space the columns within the excel file so the user no longer has to do this
# themselves this is done through the openpyxl library
def delete_rows(sheet):
    # iterate the sheet by rows
    for row in sheet.iter_rows():

        # all() return False if all of the row value is None
        if not all(cell.value for cell in row):
            # detele the empty row
            sheet.delete_rows(row[0].row, 1)

            # recursively call the remove() with modified sheet data
            delete_rows(sheet)
            return


def save_attachments(subject, messages):
    check = 0
    for message in messages:
        if message.Subject == subject:
            attachments = message.Attachments
            attachment = attachments.Item(2)
            attachment.SaveAsFile(os.path.join(path, str(attachment)))
            if message.Subject == subject and message.Unread:
                message.Unread = False
            xfile = openpyxl.load_workbook("ProspectorPatrons.xlsx")
            sheet = xfile["Sheet1"]
            delete_rows(sheet)
            for row in sheet.iter_rows():
                sheet.delete_rows(row[0].row, 1)
                break
            xfile.save("ProspectorPatrons.xlsx")
            convert_xlsx_csv()
            check = 1
            break
    if check == 1:
        print("Email found and converted successfully")
    else:
        print("Correct email not found")
        return -1


def change_filename(filename1, filename2):
    shutil.move(filename1, xlsx_file)
    shutil.move(filename2, csv_file)


def check_csv(csv_file):
    with open(csv_file) as csvfile:
        fileread = csv.reader(csvfile)
        for row in fileread:
            if row[0][-4:] == "2341":
                continue
            else:
                print("error in csv file")
                return -1
        print("File is correct")


path = os.path.expanduser("C:\\Users\\garrettthompson_a\\Downloads\\Prospector\\ProspectorAutomatedProject")
path2 = "Y:\\LB\\SharedSpace\\Systems\\Discovery\\Prospector\\" #Patron Records\\
today = datetime.date.today()
outlook = win32com.client.Dispatch('outlook.application').GetNamespace("MAPI").Folders
folder = outlook(1)
inbox = folder.Folders("Inbox")
messages = inbox.Items
messages.Sort("[ReceivedTime]", True)
date = datetime.date.today().strftime('%Y-%m-%d')
xlsx_file = path2 + "minespatrons" + date + ".xlsx"
csv_file = path2 + "minespatrons" + date + ".csv"

save_attachments("ProspectorPatrons", messages)
change_filename("ProspectorPatrons.xlsx", "ProspectorPatrons.csv")
check_csv(csv_file)
